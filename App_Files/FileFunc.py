from openpyxl import load_workbook
import os
import shutil

class File_Manager():
    def Absorption(num ,file_list, name = None):
        # 추출할 파일들의 경로 설정
        path = os.getcwd()+"/media/file/자료 취합/취합 파일/"
        colonm = ['A','B','C','D','E','F','G','H','I'] 
        results = []

        # 파일 리스트의 각 파일에서 데이터를 추출해서 리스트로 통합 저장
        for file_name_raw in file_list:
            file_name = path + file_name_raw
            wb = load_workbook(filename=file_name, data_only=True)
            ws = wb.active
        for i in range(3,ws.max_row):
                result = []
                if ws['A'+str(i)].value:
                    for j in colonm:         
                        result.append(ws[j+str(i)].value)
                else:
                    break
                results.append(result)
        
        guide = os.getcwd()+"/media/file/공지사항/admin/자료취합_양식_파일.xlsx"
        wb = load_workbook(filename=guide)
        ws = wb.active
        count_R = 3
        for line in results:
            count_C = 0
            for cell in line:
                ws[colonm[count_C]+str(count_R)].value = cell
                count_C += 1
            count_R += 1

        if name :
            wb.save(os.getcwd()+"/media/file/자료 취합/통합 파일/"+str(name)+".xlsx")
        else:
            wb.save(os.getcwd()+"/media/file/자료 취합/통합 파일/"+"통합 파일_"+str(num)+".xlsx")

    def CopyAndMove(src):
        src = src
        dst = os.getcwd()+"/media/file/자료 취합/취합 파일/"
        shutil.copy(src, dst)


'''
import os
# 설정한 경로 안의 파일들의 이름들을 리스트로 저장
file_list = os.listdir(path)
        
'''